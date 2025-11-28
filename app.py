# -*- coding: utf-8 -*-
import os
import json
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import from_excel

# ======================
# 配置管理
# ======================
CONFIG_FILE = "config.json"

def load_config():
    """加载配置文件，缺失字段自动填充默认值"""
    if not os.path.exists(CONFIG_FILE):
        default_config = {
            "excel_file": "卖货登记.xlsx",
            "sheet_name": "销售记录",
            "data_start_row": 2,
            "data_end_row": 999,
            "summary_row": 1000
        }
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(default_config, f, ensure_ascii=False, indent=2)
        print(f"✅ 默认配置已生成: {CONFIG_FILE}")
        return default_config
    
    with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
        config = json.load(f)
    
    required = ["excel_file", "sheet_name", "data_start_row", "data_end_row", "summary_row"]
    default = {
        "excel_file": "卖货登记.xlsx",
        "sheet_name": "销售记录",
        "data_start_row": 2,
        "data_end_row": 999,
        "summary_row": 1000
    }
    for key in required:
        if key not in config:
            print(f"⚠️ 配置缺失字段: {key} → 使用默认值: {default[key]}")
            config[key] = default[key]
    
    return config

CONFIG = load_config()
EXCEL_FILE = CONFIG["excel_file"]
SHEET_NAME = CONFIG["sheet_name"]
DATA_START_ROW = CONFIG["data_start_row"]
DATA_END_ROW = CONFIG["data_end_row"]
SUMMARY_ROW = CONFIG["summary_row"]

# ======================
# 工具函数
# ======================
def get_today():
    return datetime.now().strftime("%Y年%m月%d日")

def _init_sheet_structure(ws):
    """初始化工作表结构"""
    ws.delete_rows(1, ws.max_row)
    headers = ["日期", "货名", "克重", "成本单价", "成本总价",
               "平台", "货源", "卖价", "退款前利润", "退款金额", "退款后利润"]
    ws.append(headers)
    for _ in range(DATA_END_ROW - DATA_START_ROW + 1):
        ws.append([""] * 11)
    ws.cell(row=SUMMARY_ROW, column=1, value="总计")
    ws.cell(row=SUMMARY_ROW, column=5, value=f"=SUM(E{DATA_START_ROW}:E{DATA_END_ROW})")
    ws.cell(row=SUMMARY_ROW, column=9, value=f"=SUM(I{DATA_START_ROW}:I{DATA_END_ROW})")
    ws.cell(row=SUMMARY_ROW, column=11, value=f"=SUM(K{DATA_START_ROW}:K{DATA_END_ROW})")

def safe_load_workbook(filename, data_only=False):
    """安全加载工作簿（支持 data_only 模式）"""
    if not os.path.exists(filename):
        init_template(filename, SHEET_NAME)
    wb = load_workbook(filename, data_only=data_only)
    if SHEET_NAME not in wb.sheetnames:
        print(f"⚠️ 工作表 '{SHEET_NAME}' 不存在，正在创建...")
        ws = wb.create_sheet(SHEET_NAME)
        _init_sheet_structure(ws)
        wb.save(filename)
        print(f"✅ 工作表 '{SHEET_NAME}' 已创建")
    return wb

def init_template(filename, sheet_name):
    """初始化Excel模板"""
    print("ℹ️ 首次运行，正在创建Excel模板...")
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(sheet_name)
    _init_sheet_structure(ws)
    wb.save(filename)
    print(f"✅ 模板已创建: {filename}")

def find_insert_row(ws):
    """在数据区内查找第一个空行"""
    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        if ws.cell(row=row, column=1).value is None:
            return row
    return None

def format_cell_value(val):
    """将单元格值格式化为可读字符串"""
    if val is None:
        return ""
    elif isinstance(val, datetime):
        return val.strftime("%Y年%m月%d日")
    elif isinstance(val, int) and val > 30000:  # Excel日期序列号
        try:
            dt = from_excel(val)
            return dt.strftime("%Y年%m月%d日")
        except:
            return str(val)
    elif isinstance(val, float):
        return f"{val:.2f}"
    else:
        return str(val)

# ======================
# 修复对齐问题的核心函数
# ======================
def print_table(headers, rows):
    """美观打印表格（动态计算列宽）"""
    # 计算每列最大宽度
    col_widths = [len(str(h)) for h in headers]  # 初始为表头长度
    
    # 更新为数据中的最大长度
    for row in rows:
        for i, val in enumerate(row):
            val_str = str(val)
            if len(val_str) > col_widths[i]:
                col_widths[i] = len(val_str)
    
    # 确保最小宽度（避免过短）
    for i in range(len(col_widths)):
        if col_widths[i] < 4:
            col_widths[i] = 4
    
    # 打印表头
    header_line = " | ".join([f"{headers[i]:<{col_widths[i]}}" for i in range(len(headers))])
    print("=" * (len(header_line) + 2))
    print(header_line)
    print("-" * (len(header_line) + 2))
    
    # 打印数据行
    for row in rows:
        data_line = " | ".join([f"{str(row[i]):<{col_widths[i]}}" for i in range(len(row))])
        print(data_line)
    print("=" * (len(header_line) + 2))

# ======================
# 核心功能
# ======================
def add_record(excel_file, sheet_name):
    """新增销售记录（写入公式）"""
    print("\n【新增销售记录】")
    try:
        goods = input("货名: ").strip()
        weight = float(input("克重 (纯数字): "))
        cost = float(input("成本单价 (纯数字): "))
        platform = input("平台: ").strip()
        source = input("货源: ").strip()
        sell_price = float(input("卖价 (纯数字): "))
    except ValueError:
        print("❌ 输入错误！请确保克重、成本单价、卖价为数字")
        return

    total_cost = weight * cost
    profit_before = sell_price - total_cost

    wb = safe_load_workbook(excel_file, data_only=False)
    ws = wb[sheet_name]
    insert_row = find_insert_row(ws)
    if insert_row is None:
        print(f"❌ 数据区已满（最多 {DATA_END_ROW - DATA_START_ROW + 1} 条记录）！")
        return

    data = [
        get_today(), goods, weight, cost,
        f"=C{insert_row}*D{insert_row}",
        platform, source, sell_price,
        f"=H{insert_row}-E{insert_row}",
        "",
        f"=IF(J{insert_row}=\"\", MAX(0,H{insert_row}-E{insert_row}), MAX(0,H{insert_row}-E{insert_row}-J{insert_row}))"
    ]
    
    for col_idx, value in enumerate(data, start=1):
        ws.cell(row=insert_row, column=col_idx, value=value)
    wb.save(excel_file)
    
    # 回显（使用计算后的值）
    display_values = [
        get_today(), goods, f"{weight:.2f}", f"{cost:.2f}", f"{total_cost:.2f}",
        platform, source, f"{sell_price:.2f}", f"{profit_before:.2f}", "", f"{max(0, profit_before):.2f}"
    ]
    headers = ["日期", "货名", "克重", "成本单价", "成本总价",
               "平台", "货源", "卖价", "退款前利润", "退款金额", "退款后利润"]
    print("\n✅ 记录已成功添加！完整数据如下：")
    print_table(headers, [display_values])

def search_by_weight(target_weight, excel_file, sheet_name):
    """按克重搜索记录（使用 data_only=True 读取真实值）"""
    wb = safe_load_workbook(excel_file, data_only=True)
    ws = wb[sheet_name]
    matches = []
    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        cell_value = ws.cell(row=row, column=3).value  # C列：克重
        if cell_value is not None and isinstance(cell_value, (int, float)) and abs(cell_value - target_weight) < 1e-5:
            data = []
            for col in range(1, 12):
                raw_val = ws.cell(row=row, column=col).value
                formatted_val = format_cell_value(raw_val)
                data.append(formatted_val)
            matches.append((row, data))
    return matches

def process_refund(excel_file, sheet_name):
    """处理退款（仅更新J列）"""
    print("\n【处理退款】")
    print("🔍 请输入克重（必须输入，纯数字，如：17.68）")
    
    while True:
        weight_input = input("克重: ").strip()
        if not weight_input:
            print("❌ 克重不能为空！请重新输入")
            continue
        try:
            weight_val = float(weight_input)
            break
        except ValueError:
            print("❌ 克重必须是数字！请重新输入")
    
    matches = search_by_weight(weight_val, excel_file, sheet_name)
    
    if not matches:
        print(f"❌ 未找到克重 {weight_val} 的记录")
        return
    
    headers = ["日期", "货名", "克重", "成本单价", "成本总价",
               "平台", "货源", "卖价", "退款前利润", "退款金额", "退款后利润"]
    
    # 构建数据行（只包含匹配的记录）
    data_rows = []
    for i, (row_idx, data) in enumerate(matches):
        data_rows.append([f"{i+1}", f"行{row_idx}"] + data)
    
    # 打印表格（使用新对齐函数）
    print(f"\n🔍 找到 {len(matches)} 条克重 {weight_val} 的记录，请选择：")
    print_table(["序号", "行号"] + headers, data_rows)
    
    try:
        choice = int(input("选择序号: ")) - 1
        if not (0 <= choice < len(matches)):
            print("❌ 无效序号")
            return
        row_num = matches[choice][0]
    except ValueError:
        print("❌ 请输入有效数字")
        return
    
    try:
        refund = float(input("\n退款金额 (纯数字): "))
    except ValueError:
        print("❌ 退款金额必须为数字")
        return

    # 仅更新J列（第10列），使用普通模式（保留公式）
    wb = safe_load_workbook(excel_file, data_only=False)
    ws = wb[sheet_name]
    ws.cell(row=row_num, column=10, value=refund)
    wb.save(excel_file)
    
    print("✅ 退款金额已更新！")
    print(f"ℹ️ K{row_num}（退款后利润）将由公式自动计算")

# ======================
# 配置修改功能
# ======================
def modify_config():
    """修改配置文件（运行时交互式修改）"""
    print("\n" + "="*50)
    print("       修改配置")
    print("="*50)
    
    print("当前配置:")
    print(f"1. 文件名: {CONFIG['excel_file']}")
    print(f"2. Sheet名: {CONFIG['sheet_name']}")
    print(f"3. 数据区开始行: {CONFIG['data_start_row']}")
    print(f"4. 数据区结束行: {CONFIG['data_end_row']}")
    print(f"5. 统计行: {CONFIG['summary_row']}")
    print("6. 返回主菜单")
    
    choice = input("请选择要修改的配置项 (1-6): ").strip()
    
    if choice == "1":
        new_name = input("请输入新文件名 (如: 黄金销售台账.xlsx): ").strip()
        if new_name:
            CONFIG["excel_file"] = new_name
            print(f"✅ 文件名已更新为: {new_name}")
    
    elif choice == "2":
        new_sheet = input("请输入新Sheet名 (如: Daily Sales): ").strip()
        if new_sheet:
            CONFIG["sheet_name"] = new_sheet
            print(f"✅ Sheet名已更新为: {new_sheet}")
    
    elif choice == "3":
        try:
            new_start = int(input(f"请输入新开始行 (当前: {CONFIG['data_start_row']}): "))
            if new_start >= 1:
                CONFIG["data_start_row"] = new_start
                print(f"✅ 数据区开始行已更新为: {new_start}")
            else:
                print("❌ 行号必须 ≥ 1")
        except ValueError:
            print("❌ 请输入有效数字")
    
    elif choice == "4":
        try:
            new_end = int(input(f"请输入新结束行 (当前: {CONFIG['data_end_row']}): "))
            if new_end > CONFIG["data_start_row"]:
                CONFIG["data_end_row"] = new_end
                print(f"✅ 数据区结束行已更新为: {new_end}")
            else:
                print(f"❌ 结束行必须 > 开始行 ({CONFIG['data_start_row']})")
        except ValueError:
            print("❌ 请输入有效数字")
    
    elif choice == "5":
        try:
            new_summary = int(input(f"请输入新统计行 (当前: {CONFIG['summary_row']}): "))
            if new_summary > CONFIG["data_end_row"]:
                CONFIG["summary_row"] = new_summary
                print(f"✅ 统计行已更新为: {new_summary}")
            else:
                print(f"❌ 统计行必须 > 数据区结束行 ({CONFIG['data_end_row']})")
        except ValueError:
            print("❌ 请输入有效数字")
    
    elif choice == "6":
        print("↩️ 返回主菜单")
        return
    
    else:
        print("❌ 无效选项")
        return
    
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(CONFIG, f, ensure_ascii=False, indent=2)
    print(f"✅ 配置已保存到: {CONFIG_FILE}")

# ======================
# 主程序
# ======================
def main():
    while True:
        print("\n" + "="*50)
        print("       卖货登记助手")
        print("="*50)
        print("1. 新增销售记录")
        print("2. 处理退款")
        print("3. 修改配置")
        print("4. 退出")
        choice = input("请选择操作: ").strip()
        
        if choice == "1":
            add_record(EXCEL_FILE, SHEET_NAME)
        elif choice == "2":
            process_refund(EXCEL_FILE, SHEET_NAME)
        elif choice == "3":
            modify_config()
        elif choice == "4":
            print("👋 再见！")
            break
        else:
            print("❌ 无效选项，请重新选择")

if __name__ == "__main__":
    main()
